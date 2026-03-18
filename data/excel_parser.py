"""
excel_parser.py
================
Excel 解析公共模块
提取 basic/supplementary/finetune 处理器中的公共代码
"""

import os
from typing import Dict, Optional
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openai import OpenAI

from config.constants import FACTOR_ROWS, REFINE_SYSTEM_PROMPT


# ==========================================
# DeepSeek 客户端
# ==========================================

def get_deepseek_client() -> Optional[OpenAI]:
    """
    获取 DeepSeek 客户端

    Returns:
        Optional[OpenAI]: DeepSeek 客户端，API Key 未配置时返回 None
    """
    api_key = os.getenv("DEEPSEEK_API_KEY") or st.secrets.get("DEEPSEEK_API_KEY", "")
    if not api_key:
        return None
    return OpenAI(api_key=api_key, base_url="https://api.deepseek.com")


def refine_text(raw_text: str, client: Optional[OpenAI]) -> str:
    """
    调用 DeepSeek 将关键词/短语还原为连贯描述
    如果 API 不可用则返回原始文本

    Args:
        raw_text: 原始文本
        client: DeepSeek 客户端

    Returns:
        str: 还原后的连贯文本
    """
    if not client or not raw_text.strip():
        return raw_text.strip()
    try:
        resp = client.chat.completions.create(
            model="deepseek-chat",
            temperature=0,
            messages=[
                {"role": "system", "content": REFINE_SYSTEM_PROMPT},
                {"role": "user", "content": raw_text}
            ]
        )
        result = resp.choices[0].message.content.strip()
        return result if result else raw_text.strip()
    except Exception as e:
        print(f"[WARN] DeepSeek 文本还原失败，使用原始文本: {e}")
        return raw_text.strip()


# ==========================================
# 单元格安全读取
# ==========================================

def cell(ws, coord: str, default="") -> str:
    """
    安全读取单元格，返回字符串

    Args:
        ws: 工作表对象
        coord: 单元格坐标（如 "A1"）
        default: 默认值

    Returns:
        str: 单元格内容
    """
    val = ws[coord].value
    if val is None:
        return default
    return str(val).strip()


def cell_int(ws, coord: str, default: int = 0) -> int:
    """
    安全读取单元格，返回整数（用于分数）

    Args:
        ws: 工作表对象
        coord: 单元格坐标
        default: 默认值

    Returns:
        int: 单元格整数值
    """
    val = ws[coord].value
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


# ==========================================
# Sheet 解析（提取原始数据）
# ==========================================

def parse_sheet_raw(ws, client: Optional[OpenAI]) -> Optional[Dict]:
    """
    解析单个 Sheet，返回包含 text / master_comment / scores 的原始数据 Dict

    Args:
        ws: 工作表对象
        client: DeepSeek 客户端

    Returns:
        Optional[Dict]: 解析后的数据字典，内容为空时返回 None
    """
    # 1. 组装 raw_text（B1 + C4~C9）
    raw_parts = [cell(ws, "B1")]
    for _, row_num in FACTOR_ROWS:
        raw_parts.append(cell(ws, f"C{row_num}"))
    raw_text = "，".join([p for p in raw_parts if p])

    if not raw_text.strip():
        return None

    # 2. DeepSeek 文本还原
    text = refine_text(raw_text, client)

    # 3. 总评
    master_comment = cell(ws, "B2")

    # 4. 六因子评分
    scores = {}
    for factor_name, row_num in FACTOR_ROWS:
        scores[factor_name] = {
            "score": cell_int(ws, f"B{row_num}"),
            "comment": cell(ws, f"C{row_num}"),
            "suggestion": cell(ws, f"D{row_num}"),
        }

    return {
        "text": text,
        "master_comment": master_comment,
        "scores": scores,
    }