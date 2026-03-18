"""
basic_case_processor.py
========================
基础判例处理模块
"""

import time
from typing import List, Dict
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook

from data.excel_parser import get_deepseek_client, parse_sheet_raw


# ==========================================
# 公开接口
# ==========================================

def basic_case_process(uploaded_file) -> List[Dict]:
    """
    处理上传的 Excel 文件，返回基础判例列表

    Args:
        uploaded_file: Streamlit UploadedFile 对象（.xlsx / .xls）

    Returns:
        List[Dict] — 每个 Dict 为一条判例，格式与 app 中 basic_cases 一致。
        解析失败时返回空列表。
    """
    try:
        # 读取 Excel
        uploaded_file.seek(0)
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    except Exception as e:
        print(f"[ERROR] 无法读取 Excel 文件: {e}")
        st.error(f"文件读取失败: {e}")
        return []

    # 初始化 DeepSeek 客户端
    client = get_deepseek_client()
    if not client:
        st.warning("⚠️ 未配置 DeepSeek API Key，将使用原始关键词拼接作为判例文本。")

    cases = []
    total_sheets = len(wb.sheetnames)

    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        try:
            case = parse_sheet_raw(ws, client)
            if case:
                case["tags"] = "批量导入"
                case["created_at"] = time.strftime("%Y-%m-%d")
                cases.append(case)
                print(f"[INFO] Sheet [{idx}/{total_sheets}] '{sheet_name}' → 解析成功")
            else:
                print(f"[WARN] Sheet [{idx}/{total_sheets}] '{sheet_name}' → 内容为空，跳过")
        except Exception as e:
            print(f"[ERROR] Sheet [{idx}/{total_sheets}] '{sheet_name}' → 解析失败: {e}")
            continue

    print(f"[INFO] 基础判例批量导入完成: {len(cases)}/{total_sheets} 条成功")
    return cases