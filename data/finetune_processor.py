"""
finetune_processor.py
======================
微调数据处理模块（Alpaca 格式）
"""

import json
from typing import List, Dict
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook

from data.excel_parser import get_deepseek_client, parse_sheet_raw


def _build_output_text(scores: Dict) -> str:
    """
    将 scores 转成 Alpaca 的 output 文本格式

    """
    outputs = []

    for k, v in scores.items():
        if not isinstance(v, dict):
            continue

        score = v.get("score")
        comment = v.get("comment", "")

        if score is None:
            continue

        line = f"{k}{score}分，依据是{comment}"
        outputs.append(line)

    return "。".join(outputs)


# ==========================================
# Alpaca 格式组装
# ==========================================

def _build_alpaca_entry(case_data: Dict, sys_prompt: str, user_tpl: str) -> Dict:
    """
    将一条判例数据组装为 Alpaca 微调格式

    Args:
        case_data: 包含 text / scores 的 Dict
        sys_prompt: 系统提示词
        user_tpl: 用户提示词模板（含占位符）

    Returns:
        {
            "instruction": "...",
            "input": "...",
            "output": "..."
        }
    """
    text = case_data.get("text", "").strip()
    scores = case_data.get("scores", {})

    user_content = (
        user_tpl
        .replace("{product_desc}", text)
        .replace("{context_text}", "")
        .replace("{basic_case_text}", "")
        .replace("{case_text}", "")
    ).strip()

    # 如果 user_tpl 为空，就直接用原文本作为 input
    if not user_content:
        user_content = text

    output_text = _build_output_text(scores)

    return {
        "instruction": sys_prompt.strip(),
        "input": user_content,
        "output": output_text
    }


# ==========================================
# 公开接口
# ==========================================

def finetune_data_process(uploaded_file) -> List[Dict]:
    """
    处理上传的 Excel 文件，返回 Alpaca 格式的微调数据列表

    每条数据可直接以 JSON 行写入 .jsonl 微调文件

    Args:
        uploaded_file: Streamlit UploadedFile 对象（.xlsx / .xls）

    Returns:
        List[Dict] — 每个 Dict 为一条 Alpaca 训练样本。
        解析失败时返回空列表。
    """
    # --- 读取 Excel ---
    try:
        uploaded_file.seek(0)
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    except Exception as e:
        print(f"[ERROR] 无法读取 Excel 文件: {e}")
        st.error(f"文件读取失败: {e}")
        return []

    # --- 获取 Prompt 配置 ---
    prompt_cfg = st.session_state.get("prompt_config", {})
    sys_prompt = prompt_cfg.get("system_template", "")
    user_tpl = prompt_cfg.get("user_template", "")

    if not sys_prompt:
        st.warning("⚠️ 系统提示词为空，微调数据中 instruction 字段将为空字符串。")
    if not user_tpl:
        st.warning("⚠️ 用户提示词模板为空，微调数据中 input 字段将直接使用原始文本。")

    # --- 初始化 DeepSeek ---
    client = get_deepseek_client()
    if not client:
        st.warning("⚠️ 未配置 DeepSeek API Key，将使用原始关键词拼接作为训练文本。")

    # --- 逐 Sheet 解析并组装 ---
    entries = []
    total_sheets = len(wb.sheetnames)

    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        try:
            case_data = parse_sheet_raw(ws, client)
            if case_data:
                entry = _build_alpaca_entry(case_data, sys_prompt, user_tpl)
                entries.append(entry)
                print(f"[INFO] Sheet [{idx}/{total_sheets}] '{sheet_name}' → Alpaca 样本生成成功")
            else:
                print(f"[WARN] Sheet [{idx}/{total_sheets}] '{sheet_name}' → 内容为空，跳过")
        except Exception as e:
            print(f"[ERROR] Sheet [{idx}/{total_sheets}] '{sheet_name}' → 处理失败: {e}")
            continue

    print(f"[INFO] 微调数据处理完成: {len(entries)}/{total_sheets} 条成功")
    return entries
